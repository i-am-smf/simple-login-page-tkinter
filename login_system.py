from tkinter import *
from tkinter import messagebox
from PIL import ImageTk,Image
import openpyxl

main=Tk()
main.title("Main page")
main.geometry("700x600")
main.iconname('logo.ico')
main.withdraw()
signinimg=""
loginimg=""
openeyeimg=""
closeeyeimg=""

try:
    signinimg=ImageTk.PhotoImage((Image.open(f"./data/images/signin.png")).resize((50,50)))
    loginimg=ImageTk.PhotoImage((Image.open(f"./data/images/logo.png")).resize((70,50)))
    openeyeimg=ImageTk.PhotoImage((Image.open(f"./data/images/openeye.png")).resize((20,20)))
    closeeyeimg=ImageTk.PhotoImage((Image.open(f"./data/images/closeeye.png")).resize((20,20)))
except:
    messagebox.showerror(title="File not found",message="Icon Images Not found \n check images folder and ensure the files \n [ signin.png, login.png, openeye.png, closeeye.png ]")    
    quit()

try:
    wb:openpyxl.Workbook=openpyxl.load_workbook("./data/accounts.xlsx")
    ws=wb['login']
    wb.close()
except:
    wb=openpyxl.Workbook()
    wb.remove_sheet(wb.active)
    wb.create_sheet('login')
    ws=wb['login']
    l=["ACCOUNT HOLDER NAME","MOBILE","USERNAME","PASSWORD"]
    ws.append(l)
    wb.save("./data/accounts.xlsx")
    wb.close()

class config:

    def on_closing(self):
        main.destroy()

    def hover(self,button):
        button:Button=button
        button.bind("<Enter>", lambda event, button=button: self.on_enter(button))
        button.bind("<Leave>", lambda event, button=button: self.on_leave(button))

    def entryerror(self,entry):
        entry:Entry=entry
        entry.config(background="#de645d")
        entry.bind("<Enter>", lambda event, widget=entry: self.on_enter_entry(entry))

    def on_enter_entry(self,entry):
        entry:Entry=entry
        entry.config(background="white")

    def on_enter(self,button):
        button['background'] = 'SystemButtonFace'

    def on_leave(self,button):
        button['background'] = 'azure3'

    def popupboxmsg(self,msg):
        messagebox.showerror(title="Invalid Input",message=msg)
    def showpass(self,entry:Entry,label:Label):
        
        if entry.cget("show")=='':
            entry.config(show="●")
            label.config(image=closeeyeimg)
        else:
            entry.config(show="")
            label.config(image=openeyeimg)

class signup:
    def __init__(self) -> None:
        self.signinpage=Toplevel(main,background='azure3')
        self.signinpage.title("Login Page")
        self.signinpage.geometry("700x600")
        self.signinpage.resizable(False,False)

        signinimglabel=Label(self.signinpage,image=signinimg,background="azure3")
        signinimglabel.place(relx=0.5, rely=0.2, anchor=CENTER)

        accountholderlabel=Label(self.signinpage,text="Account Holder Name : ",font=("arial",15),background="azure3")
        accountholderlabel.place(relx=0.276, rely=0.35, anchor=CENTER)
        
        self.accountholderentry=Entry(self.signinpage,width=40)
        self.accountholderentry.place(relx=0.6, rely=0.35, anchor=CENTER)
        
        mobilelabel=Label(self.signinpage,text="Mobile Number : ",font=("arial",15),background="azure3")
        mobilelabel.place(relx=0.233, rely=0.42, anchor=CENTER)
        
        self.mobileentry=Entry(self.signinpage,width=40)
        self.mobileentry.place(relx=0.6, rely=0.42, anchor=CENTER)
        
        usernamelabel=Label(self.signinpage,text="Username : ",font=("arial",15),background="azure3")
        usernamelabel.place(relx=0.203, rely=0.51, anchor=CENTER)

        self.usernameentry=Entry(self.signinpage,width=40)
        self.usernameentry.place(relx=0.6, rely=0.51, anchor=CENTER)

        passwordlabel=Label(self.signinpage,text="Password : ",font=("arial",15),background="azure3")
        passwordlabel.place(relx=0.203, rely=0.58, anchor=CENTER)

        self.passwordentry=Entry(self.signinpage,width=40,show="●")
        self.passwordentry.place(relx=0.6, rely=0.58, anchor=CENTER)

        passwordrelabel=Label(self.signinpage,text="Re-Enter Password : ",font=("arial",15),background="azure3")
        passwordrelabel.place(relx=0.26, rely=0.65, anchor=CENTER)

        self.passwordreentry=Entry(self.signinpage,width=40,show="●")
        self.passwordreentry.place(relx=0.6, rely=0.65, anchor=CENTER)

        self.visiblebutton=Button(self.signinpage,image=closeeyeimg,background="azure3",command=lambda : [cf.showpass(entry=self.passwordentry,label=self.visiblebutton),cf.showpass(entry=self.passwordreentry,label=self.visiblebutton)])
        self.visiblebutton.place(relx=0.85, rely=0.61,anchor=CENTER)

        self.cancelbutton=Button(self.signinpage,text="Canel",width=9,height=2,background="azure3",command=cf.on_closing)
        self.cancelbutton.place(relx=0.26,rely=0.8,anchor=CENTER)

        self.savebutton=Button(self.signinpage,text="Save",width=9,height=2,background="azure3",command=self.details)
        self.savebutton.place(relx=0.66,rely=0.8,anchor=CENTER)

        cf.hover(self.cancelbutton)
        cf.hover(self.savebutton)

        self.signinpage.protocol("WM_DELETE_WINDOW", cf.on_closing)

    def newuser(self):
        wb:openpyxl.Workbook=openpyxl.load_workbook("./data/accounts.xlsx")
        ws=wb['login']
        details=[
            self.accountholderentry.get(),
            self.mobileentry.get(),
            self.usernameentry.get(),
            self.passwordentry.get()
        ]
        ws.append(details)
        wb.save("./data/accounts.xlsx")
        wb.close()
        self.signinpage.destroy()
        login()

    def details(self):

        for child in self.signinpage.winfo_children():
            if isinstance(child,Entry):
                if child.get()=='':
                    cf.entryerror(child)
                    child.focus_force()
                    cf.popupboxmsg(msg="Field Value missing ! ! !")
                    return

        if self.passwordentry.get() != self.passwordreentry.get():
            cf.entryerror(self.passwordentry)
            cf.entryerror(self.passwordreentry)
            cf.popupboxmsg(msg="Password must be same in both fields")
            return
        elif self.passwordentry.get() == self.passwordreentry.get():
            self.newuser()
    
class login:
    def __init__(self) -> None:
        
        self.loginpage=Toplevel(main,background="azure3")
        self.loginpage.title("Login Page")
        self.loginpage.geometry("300x450")
        self.loginpage.resizable(False,False)

        imagelabel=Label(self.loginpage,image=loginimg,background="azure3")
        imagelabel.place(relx=0.29, rely=0.2, anchor=CENTER)

        usernamelabel=Label(self.loginpage,text="Enter User Name",background="azure3")
        usernamelabel.place(relx=0.2, rely=0.3, anchor=W)
        
        self.usernameentry=Entry(self.loginpage,width=30)
        self.usernameentry.place(relx=0.21, rely=0.35, anchor=W)
        
        passwordlabel=Label(self.loginpage,text="Enter Password",background="azure3")
        passwordlabel.place(relx=0.2, rely=0.4, anchor=W)
        
        self.passwordentry=Entry(self.loginpage,width=30,show="●")
        self.passwordentry.place(relx=0.21, rely=0.45, anchor=W)

        self.passwordentry.bind("<Return>",self.logincheck)

        self.visiblebutton=Button(self.loginpage,image=closeeyeimg,background="azure3",command=lambda : cf.showpass(entry=self.passwordentry,label=self.visiblebutton))
        self.visiblebutton.place(relx=0.9, rely=0.4,anchor=CENTER)
        
        self.loginbutton=Button(self.loginpage,text="Login",background="azure3",width=7,command=self.logincheck)
        self.loginbutton.place(relx=0.6, rely=0.7, anchor=W)
        
        self.signinbutton=Button(self.loginpage,text="Sign-Up",background="azure3",width=7,command=self.signuppage)
        self.signinbutton.place(relx=0.21, rely=0.7, anchor=W)

        cf.hover(self.loginbutton)
        cf.hover(self.signinbutton)

        self.loginpage.protocol("WM_DELETE_WINDOW", cf.on_closing)

    def signuppage(self):
        self.loginpage.destroy()
        signup()
    
    def logincheck(self,e=None):
        if self.usernameentry.get()=="":
            cf.popupboxmsg(msg="Please Enter the Username ! ! !")
            return
        wb:openpyxl.Workbook=openpyxl.load_workbook("data/accounts.xlsx")
        ws=wb['login']      
        state=False
        for row in ws.values:
            if row[2]==self.usernameentry.get():
                if row[3]==self.passwordentry.get():
                    main.deiconify()
                    self.loginpage.destroy()
                    return
                elif self.passwordentry.get()=="":
                    cf.entryerror(entry=self.passwordentry)
                    cf.popupboxmsg(msg="Pleae Enter the Password ! ! !")
                else:
                    cf.entryerror(entry=self.passwordentry)
                    cf.popupboxmsg(msg="Please Enter the correct Password ! ! !")
                state=True
        if state:
            pass
        else:
            messagebox.showerror(title="Username Not Found",message="Enter the correct username \nIf your you are new user please Sign-Up ")
        wb.close()

accountholderlabel=Label(main,text="Welcome to Main Page",font=("arial",40))
accountholderlabel.place(relx=0.5, rely=0.5, anchor=CENTER)

cf=config()
lg=login()
main.mainloop()